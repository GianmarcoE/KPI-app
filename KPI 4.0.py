from tkinter import *
from tkinter.ttk import *
import openpyxl

us = ['Gianmarco', 'Qui', 'Quo', 'Qua', 'Que']

length = len(us)

#data from January, for the whole 2020 month by month
month_average_sourced = []
month_average_contacted = []
average_attendancy = []

excelbase = openpyxl.load_workbook('Database.xlsx')
sheet = excelbase["Sheet1"]

for i in range (2, 13):
    month_average_sourced.append(int(sheet.cell(row=2, column=i).value))
    month_average_contacted.append(int(sheet.cell(row=3, column=i).value))
    average_attendancy.append(int(sheet.cell(row=4, column=i).value))

window = Tk()

window.geometry('500x330')

window.minsize(500, 330)

window.configure(background='white')

window.title("KPI by Gianmarco Ercolani")

lbl = Label(window, text="Month analysed:", width= 20)

lbl.configure(background='white')

lbl.grid(column=0, row=0, padx=25, pady=(25, 45))

currentmonth = Combobox(window)

currentmonth ['values']= ("January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December")

currentmonth.current(0)

currentmonth.grid(column=1, row=0, columnspan = 2, padx=25, pady=(25, 45))

class Person(object):
    """__init__() functions as the class constructor"""
    def __init__(self, name=None, sourced=None, contacted=None, days=None, sourcedperday=None, contactedperday=None):
        self.name = name
        self.sourced = sourced
        self.contacted = contacted
        self.days = days

#make a list of sourcers with attributes from class Person
personList = []

labelsourced = Label(window, text="Sourced")
labelsourced.grid(column=1, row=1, sticky='E', padx=(0, 10))
labelsourced.configure(background='white')
labelsourced = Label(window, text="Contacted")
labelsourced.grid(column=2, row=1)
labelsourced.configure(background='white')
labelsourced = Label(window, text="Days")
labelsourced.grid(column=3, row=1)
labelsourced.configure(background='white')

sourced = [length]
contacted = [length]
days = [length]

for i in range(length):
    personList.append(Person(us[i]))
    lbl = Label(window, text=f"{us[i]}", width=20)
    lbl.grid(column=0, row=i+2, pady=2, sticky='E')
    lbl.configure(background='white')
    personList[i].sourced = Entry(window,width=10)
    personList[i].sourced.grid(column=1, row=i+2, sticky='E', pady=2)
    personList[i].contacted = Entry(window,width=10)
    personList[i].contacted.grid(column=2, row=i+2, pady=2)
    personList[i].days = Entry(window,width=10)
    personList[i].days.grid(column=3, row=i+2, pady=2)

def result():

    for i in range(length):
        personList[i].sourcedperday = int(personList[i].sourced.get())/int(personList[i].days.get())
        personList[i].contactedperday = int(personList[i].contacted.get())/int(personList[i].days.get())

    monthansw = currentmonth.current()

    result = Tk()

    result.geometry('360x395')

    result.minsize(360, 395)

    result.configure(background='white')

    result.title("Results by Gianmarco Ercolani")

    for i in range(length):
        lbl = Label (result, text=(f"{us[i]} sourced (daily average): " "{:.2f}".format(personList[i].sourcedperday)), width=60)
        lbl.configure(background='white')
        lbl.grid(column=0, pady=(10, 0))
        lbl = Label (result, text=(f"{us[i]} contacted (daily average): " "{:.2f}".format(personList[i].contactedperday)), width=60)
        lbl.configure(background='white')
        lbl.grid(column=0)

    sums = 0
    sumc = 0

    for i in range(length):
        sums += int(personList[i].sourced.get())

    for i in range(length):
        sumc += int(personList[i].contacted.get())

    averagesteam = sums/length
    averagecteam = sumc/length

    month_average_sourced.append(averagesteam)
    month_average_contacted.append(averagecteam)

    lbl = Label(result, text=(f"Team sourced candidates (daily average): " "{:.2f}".format(averagesteam)), width=60)
    lbl.grid(column=0, pady=(10, 0))
    lbl = Label(result, text=(f"Team contacted candidates (daily average): " "{:.2f}\n".format(averagecteam)), width=60)
    lbl.grid(column=0)

    month_forecast = ("January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December")

    lbl = Label(result, text=(f"Forecast team sourced candidates for {month_forecast[monthansw+1]} (per week): " "{:.2f}".format((sum(month_average_sourced)/len(month_average_sourced)) * 5 * length * (average_attendancy[monthansw + 1]/100))), width=60)
    lbl.configure(background='white')
    lbl.grid(column=0)
    lbl = Label(result, text=(f"Forecast team contacted candidates for {month_forecast[monthansw+1]} (per week): " "{:.2f}".format((sum(month_average_contacted)/len(month_average_contacted)) * 5 * length * (average_attendancy[monthansw + 1]/100))) , width=60)
    lbl.configure(background='white')
    lbl.grid(column=0)

    result.mainloop()

w = Button(window, text ="Let's go!", command=result)

w.place(relx="0.43",rely="0.87")

window.mainloop()
